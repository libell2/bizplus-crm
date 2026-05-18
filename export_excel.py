#!/usr/bin/env python3
"""
비즈플러스 CRM 엑셀 내보내기 스크립트
사용법: python3 export_excel.py crm_data.json [출력파일명.xlsx]
"""
import json, sys, re
from datetime import datetime, timedelta
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl 설치 필요: pip install openpyxl")
    sys.exit(1)

MONTHS = ['1월','2월','3월','4월','5월','6월','7월','8월','9월','10월','11월','12월']

def parse_date(d):
    if not d:
        return None
    if isinstance(d, str) and len(d) >= 10:
        try:
            return datetime.strptime(d[:10], '%Y-%m-%d')
        except:
            return None
    if isinstance(d, (int, float)):
        return datetime(1899, 12, 30) + timedelta(days=d)
    return None

def safe_int(v):
    try:
        return int(v)
    except:
        return None

def main():
    data_file = sys.argv[1] if len(sys.argv) > 1 else 'crm_data.json'
    today_str = datetime.now().strftime('%Y%m%d')
    out_file = sys.argv[2] if len(sys.argv) > 2 else f'비즈플러스_CRM_{today_str}.xlsx'
    template = Path(__file__).parent / 'export-template.xlsx'

    if not Path(data_file).exists():
        print(f"데이터 파일을 찾을 수 없습니다: {data_file}")
        sys.exit(1)
    if not template.exists():
        print(f"템플릿 파일을 찾을 수 없습니다: {template}")
        sys.exit(1)

    with open(data_file, encoding='utf-8') as f:
        data = json.load(f)

    wb = load_workbook(template)

    # ── 26인바운드관리: 26년 데이터 row 337부터 덮어쓰기 ──────────────
    ws = wb['26인바운드관리']

    # 기존 26년 행 비우기 (값만, 스타일 유지)
    for row in range(337, ws.max_row + 1):
        for col in range(2, 18):
            ws.cell(row=row, column=col).value = None

    inb26 = [r for r in data.get('inbound', [])
             if r.get('inflowYear') == '2026'
             or (isinstance(r.get('inflowDate'), str) and r.get('inflowDate','')[:4] == '2026')]
    inb26.sort(key=lambda r: str(r.get('inflowDate', '')))

    for i, r in enumerate(inb26):
        row = 337 + i
        inflow = parse_date(r.get('inflowDate'))
        meeting = parse_date(r.get('meetingDate'))
        month = MONTHS[inflow.month - 1] if inflow else ''

        ws.cell(row=row, column=2).value  = month
        ws.cell(row=row, column=3).value  = inflow
        ws.cell(row=row, column=4).value  = inflow
        ws.cell(row=row, column=5).value  = meeting
        ws.cell(row=row, column=6).value  = r.get('status', '')
        ws.cell(row=row, column=7).value  = r.get('name', '')
        ws.cell(row=row, column=8).value  = safe_int(r.get('employees'))
        ws.cell(row=row, column=9).value  = r.get('region', '')
        ws.cell(row=row, column=10).value = r.get('service', '')
        ws.cell(row=row, column=11).value = r.get('detail', '')
        ws.cell(row=row, column=13).value = r.get('contactName', '')
        ws.cell(row=row, column=14).value = r.get('phone', '')
        ws.cell(row=row, column=15).value = r.get('email', '')
        ws.cell(row=row, column=16).value = r.get('owner', '')

    print(f"  인바운드 26년: {len(inb26)}건 기록")

    # ── 26아웃바운드관리: row 9부터 덮어쓰기 ─────────────────────────
    ws2 = wb['26아웃바운드관리']

    for row in range(9, ws2.max_row + 1):
        for col in range(2, 13):
            ws2.cell(row=row, column=col).value = None

    outbound = data.get('outbound', [])
    for i, r in enumerate(outbound):
        row = 9 + i
        last_upd = None
        if r.get('lastUpdate') and r.get('lastUpdate') != '미컨택':
            last_upd = parse_date(r.get('lastUpdate'))
        meeting = parse_date(r.get('meetingDate'))
        emp = r.get('employees')

        ws2.cell(row=row, column=2).value  = r.get('name', '')
        ws2.cell(row=row, column=3).value  = safe_int(emp) if emp != '-' else None
        ws2.cell(row=row, column=4).value  = r.get('region', '')
        ws2.cell(row=row, column=5).value  = last_upd
        ws2.cell(row=row, column=6).value  = meeting
        ws2.cell(row=row, column=7).value  = r.get('service', '')
        ws2.cell(row=row, column=8).value  = r.get('detail', '')
        ws2.cell(row=row, column=9).value  = r.get('contactName', '')
        ws2.cell(row=row, column=10).value = r.get('phone', '')
        ws2.cell(row=row, column=11).value = r.get('email', '')
        ws2.cell(row=row, column=12).value = r.get('note', '')

    print(f"  아웃바운드: {len(outbound)}건 기록")

    wb.save(out_file)
    print(f"\n완료: {out_file}")

if __name__ == '__main__':
    main()
